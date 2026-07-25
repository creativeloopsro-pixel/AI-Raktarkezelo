from __future__ import annotations

import csv
import json
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path

import pdfplumber
from openpyxl import load_workbook

PARSER_VERSION = "vrp-sales-v1"

NAME_HEADERS = {
    "external name",
    "external_name",
    "megnevezes",
    "name",
    "nazov",
    "nazov polozky",
    "oznacenie",
    "oznacenie tovaru",
    "oznacenie tovaru sluzby",
    "polozka",
    "product",
    "product name",
    "termek",
    "termek neve",
}
ID_HEADERS = {
    "ean",
    "ean kod",
    "external id",
    "external product id",
    "external_id",
    "kod",
    "kod polozky",
    "kod tovaru",
    "plu",
    "product code",
    "termekkod",
}
QUANTITY_HEADERS = {
    "eladott mennyiseg",
    "mennyiseg",
    "mnozstvo",
    "predane mnozstvo",
    "quantity",
    "qty",
}
UNIT_HEADERS = {
    "egyseg",
    "jednotka",
    "merna jednotka",
    "unit",
}


class VrpReportParserError(Exception):
    code = "vrp_parser_error"


@dataclass(frozen=True)
class ParsedVrpItem:
    line_number: int
    external_product_id: str | None
    external_name: str
    quantity: Decimal
    unit: str


@dataclass(frozen=True)
class ParsedVrpError:
    line_number: int | None
    error_code: str
    message: str
    raw_row: dict[str, str]


@dataclass(frozen=True)
class ParsedVrpReport:
    items: list[ParsedVrpItem]
    errors: list[ParsedVrpError]
    canonical_items_hash: str
    parser_version: str = PARSER_VERSION


def normalize_vrp_text(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value.casefold())
    without_marks = "".join(
        character
        for character in decomposed
        if not unicodedata.combining(character)
    )
    cleaned = "".join(
        character if character.isalnum() else " " for character in without_marks
    )
    return " ".join(cleaned.split())


class VrpSalesReportParser:
    def __init__(self, *, max_rows: int = 10000):
        self.max_rows = max_rows

    def parse(self, payload: bytes, filename: str) -> ParsedVrpReport:
        suffix = Path(filename).suffix.casefold()
        if suffix == ".csv":
            rows = self._csv_rows(payload)
        elif suffix == ".xlsx":
            rows = self._xlsx_rows(payload)
        elif suffix == ".pdf":
            rows = self._pdf_rows(payload)
        else:
            raise VrpReportParserError(
                "A támogatott VRP-riportformátum: CSV, XLSX vagy szöveges PDF."
            )

        items, errors = self._parse_rows(rows)
        if not items:
            raise VrpReportParserError(
                "A riportban nem található feldolgozható termék és mennyiség."
            )
        canonical = [
            {
                "external_id": item.external_product_id or "",
                "name": normalize_vrp_text(item.external_name),
                "quantity": format(item.quantity, "f"),
                "unit": normalize_vrp_text(item.unit),
            }
            for item in items
        ]
        canonical.sort(
            key=lambda item: (
                item["external_id"],
                item["name"],
                item["quantity"],
                item["unit"],
            )
        )
        canonical_hash = sha256(
            json.dumps(
                canonical,
                ensure_ascii=False,
                separators=(",", ":"),
            ).encode("utf-8")
        ).hexdigest()
        return ParsedVrpReport(
            items=items,
            errors=errors,
            canonical_items_hash=canonical_hash,
        )

    def _csv_rows(self, payload: bytes) -> list[list[str]]:
        decoded: str | None = None
        for encoding in ("utf-8-sig", "cp1250"):
            try:
                decoded = payload.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise VrpReportParserError("A CSV kódolása nem UTF-8 vagy Windows-1250.")
        sample = decoded[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ";"
        return [
            [cell.strip() for cell in row]
            for row in csv.reader(StringIO(decoded), delimiter=delimiter)
        ]

    def _xlsx_rows(self, payload: bytes) -> list[list[str]]:
        try:
            workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
        except Exception as exc:
            raise VrpReportParserError("Az XLSX riport nem olvasható.") from exc
        try:
            sheet = workbook.active
            return [
                ["" if cell is None else str(cell).strip() for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()

    def _pdf_rows(self, payload: bytes) -> list[list[str]]:
        rows: list[list[str]] = []
        try:
            with pdfplumber.open(BytesIO(payload)) as pdf:
                for page in pdf.pages:
                    for table in page.extract_tables():
                        rows.extend(
                            [
                                ["" if cell is None else str(cell).strip() for cell in row]
                                for row in table
                                if row
                            ]
                        )
        except Exception as exc:
            raise VrpReportParserError("A PDF riport nem olvasható.") from exc
        if not rows:
            raise VrpReportParserError(
                "A PDF nem tartalmaz géppel olvasható tételtáblát; használj CSV vagy XLSX exportot."
            )
        return rows

    def _parse_rows(
        self, rows: list[list[str]]
    ) -> tuple[list[ParsedVrpItem], list[ParsedVrpError]]:
        header_index, columns = self._find_header(rows)
        items: list[ParsedVrpItem] = []
        errors: list[ParsedVrpError] = []
        data_rows = rows[header_index + 1 :]
        if len(data_rows) > self.max_rows:
            raise VrpReportParserError(
                f"A riport legfeljebb {self.max_rows} adatsort tartalmazhat."
            )

        for row_offset, row in enumerate(data_rows, start=header_index + 2):
            if not any(str(value).strip() for value in row):
                continue
            raw = {
                "external_product_id": self._cell(row, columns.get("id")),
                "external_name": self._cell(row, columns["name"]),
                "quantity": self._cell(row, columns["quantity"]),
                "unit": self._cell(row, columns.get("unit")),
            }
            name = raw["external_name"].strip()
            if not name:
                errors.append(
                    ParsedVrpError(
                        line_number=row_offset,
                        error_code="MISSING_PRODUCT_NAME",
                        message="A termék megnevezése hiányzik.",
                        raw_row=raw,
                    )
                )
                continue
            if len(name) > 255:
                errors.append(
                    ParsedVrpError(
                        line_number=row_offset,
                        error_code="PRODUCT_NAME_TOO_LONG",
                        message="A termék megnevezése túl hosszú.",
                        raw_row=raw,
                    )
                )
                continue
            try:
                quantity = self._decimal(raw["quantity"])
            except (InvalidOperation, ValueError):
                errors.append(
                    ParsedVrpError(
                        line_number=row_offset,
                        error_code="INVALID_QUANTITY",
                        message="A mennyiség nem értelmezhető számként.",
                        raw_row=raw,
                    )
                )
                continue
            if quantity == 0 or abs(quantity) > Decimal("1000000000000"):
                errors.append(
                    ParsedVrpError(
                        line_number=row_offset,
                        error_code="INVALID_QUANTITY",
                        message="A mennyiség nem lehet nulla vagy kirívóan nagy.",
                        raw_row=raw,
                    )
                )
                continue

            external_id = raw["external_product_id"].strip() or None
            if external_id is not None:
                external_id = external_id[:160]
            items.append(
                ParsedVrpItem(
                    line_number=row_offset,
                    external_product_id=external_id,
                    external_name=name,
                    quantity=quantity.quantize(Decimal("0.001")),
                    unit=(raw["unit"].strip() or "piece")[:80],
                )
            )
        return items, errors

    def _find_header(self, rows: list[list[str]]) -> tuple[int, dict[str, int]]:
        for row_index, row in enumerate(rows[:30]):
            normalized = [normalize_vrp_text(str(cell)) for cell in row]
            columns: dict[str, int] = {}
            for index, header in enumerate(normalized):
                if header in NAME_HEADERS:
                    columns["name"] = index
                elif header in ID_HEADERS:
                    columns["id"] = index
                elif header in QUANTITY_HEADERS:
                    columns["quantity"] = index
                elif header in UNIT_HEADERS:
                    columns["unit"] = index
            if "name" in columns and "quantity" in columns:
                return row_index, columns
        raise VrpReportParserError(
            "Nem található terméknév- és mennyiségoszlop a VRP-riportban."
        )

    @staticmethod
    def _cell(row: list[str], index: int | None) -> str:
        if index is None or index >= len(row):
            return ""
        return str(row[index]).strip()

    @staticmethod
    def _decimal(value: str) -> Decimal:
        normalized = value.strip().replace("\u00a0", "").replace(" ", "")
        if not normalized:
            raise ValueError
        if "," in normalized and "." in normalized:
            if normalized.rfind(",") > normalized.rfind("."):
                normalized = normalized.replace(".", "").replace(",", ".")
            else:
                normalized = normalized.replace(",", "")
        else:
            normalized = normalized.replace(",", ".")
        return Decimal(normalized)
